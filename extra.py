import pandas as pd
class extra(pd.DataFrame):
    """
    Ajoute des fonctions aux dataframes pandas.
    
    Usage : 
      import pandas as pd
      df              = pd.DataFrame()
      dfextra         = extra(df)
      dfextra.calculate_stats()
      dfextra.analyse_to_excel()
    """
        
    def give_categorical_columns(self, limite=50):
        """
        Retourne la liste des colonnes de la df qui contient moins de 'limite' modalités.
        
        Parameters : 
            limit : int, nombre de modalité limite (exclusif) que doivent avoir les colonnes retournées

        Return:
            Array of strings : each string is a column name
            
        """
        return [col for col in self.columns if len(self[col].unique()) < limite and "DT_" not in col]
    
    def histo(self, cols):
        """
        S'appuie sur la df contenue dans l'objet actuel
        Pour chaque colonne : 
            \-> Compte le nombre de modalité dans chacune des colonnes
                \-> Crée :
                    \-> un graphique en bar avec ce compte par modalité
                    \-> des graphiques de zoom s'il y a assez de modalités :
        Parameters:
            cols : array of string
                \-> chaque string est un nom de colonne du tableau
                            
        """
        import matplotlib
        import matplotlib.pyplot as plt
        matplotlib.style.use('ggplot')
        for col in cols:
            try:
                count_ = self.groupby(col)[col].count()
                count_.sort_values(inplace=True, ascending=False)
                count_.index=[" %s "%x for x in count_.index]
                
                for kind in ['bar', 'pie'][0:count_.shape[0]-1]:
                    plt.figure()                    
                    count_.plot(kind=kind, title="%s - %s valeurs"%(col,count_.shape[0]))
                    if count_.shape[0]>8:
                        plt.figure()
                        count_[1:5].plot(kind=kind, title="%s - valeurs 2 à 4 "%(col))
                        plt.figure()                    
                        count_[5:].plot(kind=kind, title="%s - valeurs 4 et plus "%(col))
                        
            except Exception as e:
                print("Pbm sur col [%s] -> %s"%(col, e))
    def get_time_col(self):
        """
        Retourne la liste des colonnes contenant "DT_" dans leur nom.
        """
        return [col for col in self.columns if "DT_" in col]

    @staticmethod
    def timeline(path, col_date, debug=False):    
        """
        Créer des graphs de séries temporelles à partir d'un csv : 
            - créer une pandas df depuis le chemin
            - trouve les colonnes catégorielles
                - pour chacune 
                    \-> crée un graphique pour chaque valeur de la colonne
                    \-> enregistre le graphique sur le disque  
        Parameters:
            path : string, chemin vers le csv de datframe
            col_date : string, colonne de date sur laquelle se baser pour le temps
            debug : boolean

        Return: None
        
        """
        import os
        table     = extra(pd.read_csv(path, delimiter=useful.find_delimiter(path)))
        timeline2(table, path, col_date, debug=False)

    @staticmethod
    def timeline2(df, path, col_date, debug=False):    
        """
        Créer des graphs de séries temporelles à partir d'un csv : 
            - créer une pandas df depuis le chemin
            - trouve les colonnes catégorielles
                - pour chacune 
                    \-> crée un graphique pour chaque valeur de la colonne
                    \-> enregistre le graphique sur le disque  
        Parameters:
            path : string, chemin vers le csv de datframe
            col_date : string, colonne de date sur laquelle se baser pour le temps
            debug : boolean

        Return: None
        
        """
        import os
        
        test2     = extra(df)
        
        modalities, binary, continuous, dates = extra_pandas.select_columns_based_on_names(test2)
        
        cat_col   = test2.give_categorical_columns()
        
        #extra_pandas.enforce_date(test2, col_date, "%d%b%Y:00:00:00", debug=True)
        index     = test2[col_date]
        nom_table = path[path.rfind("/")+1:path.rfind(".")]
        for col in cat_col:
            if debug : print("col = ", col)    
            categories = list(test2[col].value_counts().keys())
            if debug : print("categories = ", categories)
            for cat in categories : 
                test5         = test2[col]
                test5.index   = index
                test6         = pd.DataFrame(test5)
                test6["date"] = test6.index
                test6["one"]  = 1
                print("cat=", cat)
                year_mini   = 2013
                year_maxi   = 2016
                maxi1       = test6[test6[col]==cat]
                vrai_maxi   = maxi1["%s"%year_mini:"%s"%year_maxi].groupby("date").count()['one'].max() 
                maxi        = vrai_maxi * 1.05
                mini        = vrai_maxi * 0.05
                
                #for year in range(2014,2017):
                try:
                    plt.figure(figsize=(20,5))
                    n1 = test6["%s"%year_mini:"%s"%year_maxi] 
                    n2 = n1[n1[col]==cat]
                    n3 = n2.groupby("date").count()
                    n4 = n3.resample('d')
                    nom_table_utf8 = to_Str(nom_table)
                    col_utf8       = to_Str(col)
                    cat_utf8       = to_Str(cat)
                    col_date_utf8  = to_Str(col_date)
                    """
                    try    : nom_table_utf8 = to_Str(nom_table)
                    except : nom_table_utf8 = nom_table
                    try    : col_utf8       = col.decode("utf-8") 
                    except : col_utf8       = col
                        try : col_utf8       = col.decode('ascii', errors='replace')
                        except : col_utf8 = "illisible"
                    try    : cat_utf8       = cat.decode("utf-8")
                    except : 
                        try : cat_utf8       = cat.decode('ascii', errors='replace')
                        except : cat_utf8 = "illisible"
                    try    : col_date_utf8  = col_date.decode("utf-8")
                    except : col_date_utf8  = col_date
                    """    
                    title_ = "TABLE_[%s]_COLONNE_[%s]_CATEGORIE_[%s]_nombre_par_jour_par_date_de_[%s]"%(nom_table_utf8, col_utf8, cat_utf8, col_date_utf8)
                    ax = n4["one"].sum().plot(style='o', title=title_)
                    ax.set_axis_bgcolor('w')
                    ax.set_ylim(-mini,maxi)

                    ax.plot()
                    title_ = title_.replace("/", "").replace(" ", "_")
                    savedir = "./../images/table_%s/col_%s/"%(nom_table_utf8.replace("/", "").replace(" ", "_"), col_utf8.replace("/", "").replace(" ", "_"))
                    print(savedir)
                    try: 
                        os.makedirs(savedir)
                    except OSError:
                        if not os.path.isdir(savedir):
                            raise
                    plt.savefig(savedir+"%s.jpg"%title_)
                    plt.show()

                    #ax.set_ylim(maxi)
                except Exception as e:
                    print("e=", e)
                    pass
    def calculate_stats(self):
        r = {}
        df = self

        for col in self.columns:
            nb_valeur     = df[col].value_counts().shape[0]
            nb_null       = df[col].isnull().sum()
            n             = 10
            top_n         = df[col].value_counts().sort_values(ascending=False).reset_index()[:n].values
            pct_top_n     = [x[1]/df.shape[0] for x in top_n]
            clef_top_n    = [x[0] for x in top_n]
            val_top_n     = [x[1] for x in top_n]
            import numpy as np
            r[col] = {
                        "nb_valeur"          : nb_valeur      ,
                        "nb_null"            : nb_null        ,
                        "n"                  : n              ,
                        f"top_{n}"           : top_n          ,
                        f"clef_top_{n}"      : clef_top_n     ,
                        f"val_top_{n}"       : val_top_n      ,
                        f"pct_top_{n}"       : pct_top_n      ,
                    }
            subdf       = pd.DataFrame()
            col_gauche  = 0
            ligne_debut = 0
            col_droite  = col_gauche + 1
            col_pct     = col_droite + 1
            ligne_n     = ligne_debut
            to_display  = [["nb_valeur", nb_valeur], 
                           ["nb_null", nb_null ],
                           ["valeur", "nb occurence"]]
            to_display.extend(top_n)
            subdf.loc [ligne_n , col_gauche] = str(col)
            for valeur, nb_occurence in to_display:
                ligne_n    += 1
                subdf.loc [ligne_n , col_gauche] = str(valeur)
                subdf.loc [ligne_n , col_droite] = nb_occurence     
            
            ligne_n = ligne_debut +3
            
            for pct in pct_top_n :
                ligne_n +=1
                subdf.loc [ligne_n , "percent"] = pct
            
            r[col]["subdf"]    = subdf
            try:
                df_couleurs        = subdf[["percent"]].applymap(lambda x: self.get_excel_style(x))
                r[col]["df_couleurs"]    = df_couleurs

                list_couleurs      = np.reshape(df_couleurs.values, (1, df_couleurs.shape[0]))
                r[col]["couleurs"] = list(list(list_couleurs)[0])*3
            except:
                r[col]["couleurs"] = [self.get_excel_style(0) for i in range( 20 )]
            

            
        self.analyse = r
    def set_style(self):
        def get_style(name, font_color, filling_color):
            from openpyxl.styles import PatternFill
            from openpyxl.styles import Font
            from openpyxl.styles import NamedStyle 
            from openpyxl.styles import Font, Color
            import random
            import string
            filling = PatternFill(start_color = Color(filling_color), 
                                   end_color  = Color(filling_color), 
                                   fill_type  = "solid")
            
            font    = Font( color = Color(font_color),
                            b     = True)
            
            style   = NamedStyle(name = name + "_" + "".join([random.choice(string.ascii_letters) for i in range(10)]), 
                                 font  = font, 
                                 fill  = filling)
            return style
        
        from openpyxl.styles import colors
        self.red_on_yellow   = get_style("red_on_yellow", colors.RED, colors.YELLOW)
        self.red_on_white    = get_style("red_on_white", colors.RED, colors.WHITE)
        self.red_on_black    = get_style("red_on_black", colors.RED, colors.BLACK)
        self.black_on_white  = get_style("black_on_white", colors.BLACK, colors.WHITE)
        self.black_on_yellow = get_style("black_on_yellow", colors.BLACK, colors.YELLOW)
        self.white_on_green  = get_style("white_on_green", colors.WHITE, colors.GREEN)
        self.white_on_darkblue  = get_style("white_on_darkblue", colors.WHITE, colors.DARKBLUE)
        self.red_on_darkblue  = get_style("red_on_darkblue", colors.RED, colors.DARKBLUE)
    
    def get_excel_style(self, value):
        self.set_style()
        if not hasattr(self, "black_on_white"):
            self.set_style()
        try:
            value = float(value)
            
        except:
            print(f"value is not float : {value}")
            return self.black_on_white
        
        
        n               = 10
        limites         = { 0   : self.black_on_white,
                            0.3 : self.black_on_yellow,
                            0.4 : self.white_on_green,
                            0.6 : self.white_on_darkblue,
                            0.8 : self.red_on_black ,
                           }
        try:
            possible_keys = filter(lambda limite: limite < value, limites.keys())
            max_key       = max(possible_keys)
            couleur       = limites[max_key]
            return couleur
        except:
            return self.black_on_white
        
    def print_df(self, sheet, df, starting_row,starting_col, couleurs):
        """
        Ecrit le contenu d'une dataframe dans une worksheet d'un fichier excel
        Param:
            sheet : worksheet excel
            df : pandas dataFrame
            starting_row : entier, ligne de départ
            starting_col : entier, ligne de départ
            couleurs : liste de couleurs (df.shape[0] * df.shape[1])

        """
        from openpyxl.utils import get_column_letter    
        for current_row in range(df.shape[0]):
            writing_row = starting_row+current_row+1
            writing_row = str(get_column_letter(writing_row))

            for current_col in range(df.shape[1]):

                v = df.iloc[current_row,current_col]
                v = float(v) if str(type(v)) == "<class 'numpy.int64'>" else v
                v = v if v==v else ""


                writing_col = starting_col+current_col+1
                writing_col = str(int(writing_col))

                emplacement = writing_row+writing_col

                sheet[emplacement] =  v
                try:
                    sheet[emplacement].style = couleurs.pop(0)
                except Exception as e:
                    print(e)     
                    
    def analyse_to_excel(self):
        from openpyxl        import Workbook, load_workbook
        dfextra = self
        #create excel type item
        wb = Workbook()
        # select the active worksheet
        ws = wb.active
        style  = None
        base_i = 1
        base_j = 1
        previous_row = 0

        for n_eme, colonne in enumerate(dfextra.analyse.keys()):
            starting_col = (n_eme % 4) * 15
            starting_row = (n_eme // 4) *5
            print(colonne, end=" / ")        
            if True:
                df       = dfextra.analyse[colonne]["subdf"].T
                couleurs = dfextra.analyse[colonne]["couleurs"]
                self.print_df(ws, df, starting_row,starting_col, couleurs)
        import string
        import random
        output_file = "test_" + ''.join([random.choice(string.ascii_letters) for i in range(5)]) + ".xlsx"
        print(f"saving into {output_file}")
        wb.save(output_file)
                    
